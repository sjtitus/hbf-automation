"""
Loader and validator for the customer-address master list.

Reads `data/hbf-customer-shipping-addresses.xlsx` and returns a map from
`NormalizedAddress` (street + line_2 + city + state + postcode) to a
list of `CustomerEntry` (`name` from the `Name` column, `shipto_name`
derived from `AddressLine1` via the new `parse_al1` parser).

**Schema model** (per the human-maintained source file):
  - `Name`         = the **Customer** (billing entity).
  - `AddressLine1` = `<shipping-name-info> - <hbf customer number>[<sep><additional-info>]`
                     where `<sep>` is `,` and/or whitespace (comma optional).
                     For no-dash rows, the entire AL1 is the shipping name.
  - `AddressLine2` = the actual street.
  - `City`/`State`/`Postcode` = standard.

A single canonical address can hold many customers (multi-tenant sites
like federal complexes); a single Customer can have many addresses
(distributors with multiple ship-to locations); the row-level uniqueness
invariant is `(Name, normalized shipto_name, NormalizedAddress)`.

**Validation phase** runs at load time. Hard-rule violations
(missing required fields, duplicate triples) are reported and — when
`strict=True` — abort the load via `MasterValidationError`. Soft-rule
warnings (AL1 strict-parse failure, no-dash AL1, scourgify-fallback
addresses) are reported but never abort. A human-readable validation
log file is written to `<log_dir>/customer_master_validation.log` when
`log_dir` is provided.

NOTE (stage 1): the matching logic below predates the 5-field address
shape; `line_2` is captured but ignored at match time. Stage-2 makes it
a real participant.
"""

from __future__ import annotations

import logging
import re
from collections import namedtuple
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional, Union

from openpyxl import load_workbook
from rapidfuzz import fuzz

from .ship_to import (
    NormalizedAddress,
    _normalize_address,
    _normalize_address_with_status,
    _norm,
    _fmt_postcode,
)


logger = logging.getLogger(__name__)


# `name` from the Name column; `shipto_name` is the parsed shipping-name
# portion of AL1 (the pre-dash text for dash rows, full AL1 for no-dash
# rows). For non-distributors the two are typically equivalent modulo
# minor naming variations; for distributors they genuinely differ.
CustomerEntry = namedtuple('CustomerEntry', 'name shipto_name')
LookupResult = namedtuple(
    'LookupResult',
    'pairs cm_method addr_score name_method name_score',
)
# pairs:       list[(NormalizedAddress, CustomerEntry)] — match output, OR
#              best near-miss on no_match / multi_match_unresolved.
# cm_method:   one of:
#                'address_exact'                — exact address hit, single match
#                'address_fuzzy'                — fuzzy address hit, single match
#                'address_disambiguated_by_name' — address multi-hit, name picked one
#                'name_fallback'                — address missed, name found a match
#                'multi_match_unresolved'       — address multi-hit, name couldn't narrow
#                                                 (or generic consignee blocked the attempt)
#                'no_match'                     — neither tier matched
# addr_score:  address-tier score; 100 on address_exact, fuzz score on address_fuzzy,
#              best near-miss score on no_match (0 when no candidates considered).
# name_method: which name tier was used (only meaningful when name was actually used):
#                'exact'             — tier 1: verbatim equality
#                'case_insensitive'  — tier 2
#                'normalized'        — tier 3
#                'fuzzy'             — tier 4: WRatio at or above threshold
#                'tried_failed'      — name was attempted (fallback or disambiguator)
#                                      but no tier produced a unique winner
#                'n/a'               — name was not used at all (single-match address,
#                                      or generic consignee blocked the attempt)
# name_score:  100 for tiers 1–3, WRatio for tier 4, best near-miss WRatio on
#              tried_failed, 0 when name_method == 'n/a'.


# Default minimum token_set_ratio to accept a fuzzy address fallback within
# a CSZ bucket. Tunable per-call via lookup_by_address(..., fuzzy_threshold=…).
DEFAULT_FUZZY_THRESHOLD = 85

# Default minimum rapidfuzz WRatio score (0–100) to accept a fuzzy name match.
# WRatio blends ratio / partial_ratio / token_sort_ratio / token_set_ratio
# with sensible weights, so it handles both the "abbreviation vs full" case
# and the "containment" case (consignee = customer name + extra qualifiers
# like 'Commissary Whse') that strict ratio misses.
DEFAULT_NAME_FUZZY_THRESHOLD = 88

# Consignee strings that are too generic to fuzzy-match by name safely
# (would partial-match every facility of that type in the customer set).
# Comparison is against the name-normalized form (lowercase, alphanumerics
# only, whitespace collapsed).
_GENERIC_CONSIGNEE_NORMS = frozenset({
    'federal correctional institution',
})


_PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_ADDRESS_FILE = _PROJECT_ROOT / 'data' / 'hbf-customer-shipping-addresses.xlsx'


# AL1 format (per source-file convention):
#   <shipping-name-info> - <hbf customer number>[<sep><additional-info>]
# where <sep> is `,` and/or whitespace (comma optional).
# Last `\s+-\s*<digits>` wins (greedy `.+`), so internal-dash names like
# "Avery-Mitchell Corr. Inst." are preserved. Whitespace required BEFORE
# the dash so internal hyphens don't false-match.
_AL1_PARSE_RE = re.compile(r'^(.+)\s+-\s*(\d+)(?:[\s,]+(.*?))?\s*$')

# Detects the customer-number separator pattern: whitespace before a
# dash. This distinguishes a real separator from internal hyphens like
# 'GSNA-Jekyll' or 'Cash-WA Distributing' (which have NO whitespace
# before the dash). Permissive about what follows the dash so rows
# like 'Center -205997' (missing space after dash) are still recognized
# as customer-number attempts. A row WITHOUT this pattern is treated as
# a no-customer-number row regardless of internal hyphens.
_CUSTOMER_NUM_SEP_RE = re.compile(r'\s+-')


def parse_al1(al1) -> tuple[str, Optional[str], Optional[str], bool]:
    """Parse an AL1 cell.

    Returns `(shipping_name, customer_number, additional_info, parsed_ok)`:
      - dash row matching the format → all four populated, parsed_ok=True.
      - no-customer-number row (no customer-number separator pattern, OR
        internal-hyphens-only) → (al1.strip(), None, None, True). Whole
        AL1 is the shipping name; parse considered successful.
      - dash row with the separator pattern that doesn't strict-parse →
        (al1.strip(), None, None, False). Validation flags these.
      - empty / None input → ('', None, None, True).
    """
    if al1 is None:
        return '', None, None, True
    s = str(al1).strip()
    if not s:
        return '', None, None, True

    # No customer-number separator pattern → treat as no-customer-number
    # row regardless of any internal hyphens within words. The whole AL1
    # is the shipping name. (E.g. 'ABF Freight c/o PeakXpo GSNA-Jekyll
    # Island' — only dash is internal in 'GSNA-Jekyll'.)
    if not _CUSTOMER_NUM_SEP_RE.search(s):
        return s, None, None, True

    m = _AL1_PARSE_RE.match(s)
    if not m:
        return s, None, None, False

    name = m.group(1).strip()
    number = m.group(2)
    extra = (m.group(3) or '').strip() or None
    return name, number, extra, True


# ============================================================
# Validation
# ============================================================


class MasterValidationError(Exception):
    """Raised when `load_address_to_customers(strict=True)` finds at least
    one hard-rule violation in the customer master."""


@dataclass(frozen=True)
class ValidationViolation:
    row: int          # 1-indexed XLSX row (header is row 1, data starts row 2)
    rule: str         # rule identifier — one of the RULE_ keys below
    severity: str     # 'hard' | 'soft'
    message: str      # human-readable explanation
    raw: dict         # row data: {Name, AddressLine1, AddressLine2, City, State, Postcode}


# Rule identifiers — kept as constants so the report has a stable
# vocabulary and tests can refer to them.
RULE_REQUIRED_FIELDS  = 'required_fields_present'    # hard
RULE_TRIPLE_UNIQUE    = 'triple_unique'              # hard
RULE_AL1_STRICT_PARSE = 'al1_strict_parse'           # soft
RULE_NO_DASH_AL1      = 'no_dash_al1'                # soft
RULE_SCOURGIFY_FALL   = 'scourgify_fallback'         # soft

_HARD_RULES = {RULE_REQUIRED_FIELDS, RULE_TRIPLE_UNIQUE}


# Per-row record built by the loader, consumed by validation.
@dataclass(frozen=True)
class _RowRecord:
    row: int                              # 1-indexed XLSX row
    raw: dict                             # original cell values
    shipto_name: str                      # parsed AL1 shipping name
    customer_number: Optional[str]
    al1_extra: Optional[str]
    al1_parsed_ok: bool                   # AL1 strict-parse status
    address: Optional[NormalizedAddress]  # None if CSZ missing
    address_used_fallback: bool           # scourgify fell back to plain norm


def validate_master(records: list[_RowRecord]) -> list[ValidationViolation]:
    """Run all validation rules over the loaded row records and return
    the consolidated violation list. Rules are applied independently;
    each rule's violations are accumulated in input row order.
    """
    violations: list[ValidationViolation] = []
    violations.extend(_check_required_fields(records))
    violations.extend(_check_triple_unique(records))
    violations.extend(_check_al1_strict_parse(records))
    violations.extend(_check_no_dash_al1(records))
    violations.extend(_check_scourgify_fallback(records))
    violations.sort(key=lambda v: (v.row, v.rule))
    return violations


def _check_required_fields(records: list[_RowRecord]) -> list[ValidationViolation]:
    """Hard rule: Name, AddressLine1, AddressLine2, City, State, Postcode
    must all be non-empty."""
    out = []
    for r in records:
        missing = []
        for field in ('Name', 'AddressLine1', 'AddressLine2',
                      'City', 'State', 'Postcode'):
            v = r.raw.get(field)
            if v is None or (isinstance(v, str) and not v.strip()):
                missing.append(field)
        if missing:
            out.append(ValidationViolation(
                row=r.row, rule=RULE_REQUIRED_FIELDS, severity='hard',
                message=f"missing required field(s): {missing}",
                raw=r.raw,
            ))
    return out


def _check_triple_unique(records: list[_RowRecord]) -> list[ValidationViolation]:
    """Hard rule: `(Name, normalized shipto_name, NormalizedAddress)` is
    unique across all rows. Skip rows whose address couldn't be built
    (missing CSZ) — those are already flagged by required_fields."""
    seen: dict[tuple, int] = {}
    out = []
    for r in records:
        if r.address is None:
            continue
        name = (r.raw.get('Name') or '').strip()
        if not name:
            continue
        key = (name, _normalize_name(r.shipto_name), r.address)
        prev = seen.get(key)
        if prev is not None:
            out.append(ValidationViolation(
                row=r.row, rule=RULE_TRIPLE_UNIQUE, severity='hard',
                message=(f"duplicate (Name={name!r}, "
                         f"shipto_name={r.shipto_name!r}, address) — "
                         f"first seen at row {prev}"),
                raw=r.raw,
            ))
        else:
            seen[key] = r.row
    return out


def _check_al1_strict_parse(records: list[_RowRecord]) -> list[ValidationViolation]:
    """Soft rule: AL1 that uses the customer-number separator pattern
    (whitespace-dash) should strict-parse to a clean
    `<name> - <digits>[<sep><extra>]`. Flags rows that LOOK like
    they're trying to have a customer number but the format is wrong
    (Lompoc-style missing dash before number, `#`-prefix on number,
    orphan `- C` suffix, etc.). Rows with no customer-number separator
    pattern (only internal hyphens or no dash at all) are flagged by
    the no-customer-number rule instead, not here.
    """
    out = []
    for r in records:
        al1 = r.raw.get('AddressLine1')
        if al1 is None:
            continue
        s = str(al1)
        if not _CUSTOMER_NUM_SEP_RE.search(s):
            continue  # no separator pattern → not this rule's concern
        if not r.al1_parsed_ok:
            out.append(ValidationViolation(
                row=r.row, rule=RULE_AL1_STRICT_PARSE, severity='soft',
                message=(f"AL1 has a customer-number separator but "
                         f"doesn't match expected format: {al1!r}"),
                raw=r.raw,
            ))
    return out


def _check_no_dash_al1(records: list[_RowRecord]) -> list[ValidationViolation]:
    """Soft rule: AL1 with no customer-number separator pattern (no
    `\\s+-\\s` anywhere). Includes both true no-dash rows and rows
    whose only dashes are internal hyphens within words (e.g.,
    'ABF Freight c/o PeakXpo GSNA-Jekyll Island'). Flags the row for
    human review; legitimate no-customer-number rows are fine to
    ignore.
    """
    out = []
    for r in records:
        al1 = r.raw.get('AddressLine1')
        if al1 is None:
            continue
        if not _CUSTOMER_NUM_SEP_RE.search(str(al1)):
            out.append(ValidationViolation(
                row=r.row, rule=RULE_NO_DASH_AL1, severity='soft',
                message=f"AL1 has no customer number assigned: {al1!r}",
                raw=r.raw,
            ))
    return out


def _check_scourgify_fallback(records: list[_RowRecord]) -> list[ValidationViolation]:
    """Soft rule: address didn't parse via scourgify; loader fell back
    to plain whitespace-collapse + uppercase. May create asymmetric
    matching against an OCR-extracted invoice address that DOES parse."""
    out = []
    for r in records:
        if r.address_used_fallback and r.address is not None:
            al2 = r.raw.get('AddressLine2')
            out.append(ValidationViolation(
                row=r.row, rule=RULE_SCOURGIFY_FALL, severity='soft',
                message=(f"scourgify could not parse the street; using "
                         f"plain norm fallback. AL2={al2!r} → "
                         f"street={r.address.street!r}"),
                raw=r.raw,
            ))
    return out


def write_validation_report(
    violations: list[ValidationViolation],
    *,
    source_path: Path,
    total_rows: int,
    log_path: Path,
    strict: bool,
    aborted: bool = False,
) -> None:
    """Write a human-readable validation report to `log_path`. Always
    overwrites. Sections per rule, row-by-row entries, summary footer.
    """
    log_path.parent.mkdir(parents=True, exist_ok=True)

    # Group by rule, preserving sort-by-row order within each group.
    by_rule: dict[str, list[ValidationViolation]] = {}
    for v in violations:
        by_rule.setdefault(v.rule, []).append(v)

    n_hard = sum(1 for v in violations if v.severity == 'hard')
    n_soft = sum(1 for v in violations if v.severity == 'soft')

    if aborted:
        verdict = 'FAIL (aborted via --strict-master)'
    elif n_hard > 0 and strict:
        verdict = 'FAIL'
    elif n_hard > 0:
        verdict = (f'PASS with {n_hard} hard violation(s) '
                   f'(non-strict mode; would FAIL with --strict-master)')
    else:
        verdict = 'PASS'

    lines = []
    lines.append("HBF Customer Master Validation Report")
    lines.append("=" * 78)
    lines.append(f"Generated:   {datetime.now().isoformat(timespec='seconds')}")
    lines.append(f"Source:      {source_path}")
    lines.append(f"Total rows:  {total_rows}")
    lines.append(f"Mode:        {'strict (hard violations are fatal)' if strict else 'non-strict (failures logged, run continues)'}")
    lines.append("")

    rule_specs = [
        (RULE_REQUIRED_FIELDS, 'hard', 'Required fields non-empty'),
        (RULE_TRIPLE_UNIQUE,   'hard', 'Triple uniqueness (Name, shipto_name, address)'),
        (RULE_AL1_STRICT_PARSE,'soft', 'AL1 strict-parse failure'),
        (RULE_NO_DASH_AL1,     'soft', 'AL1 has no customer number assigned'),
        (RULE_SCOURGIFY_FALL,  'soft', 'Scourgify-fallback addresses'),
    ]

    lines.append("=== HARD RULES ===")
    lines.append("")
    for rule, severity, label in rule_specs:
        if severity != 'hard':
            continue
        viols = by_rule.get(rule, [])
        status = 'FAIL' if viols else 'OK'
        lines.append(f"[{status}] {label}  [rule={rule}]  ({len(viols)} violation(s))")
        for v in viols:
            lines.append(f"  row {v.row:3d}: {v.message}")
        lines.append("")

    lines.append("=== SOFT RULES (warnings) ===")
    lines.append("")
    for rule, severity, label in rule_specs:
        if severity != 'soft':
            continue
        viols = by_rule.get(rule, [])
        status = 'WARN' if viols else 'OK'
        lines.append(f"[{status}] {label}  [rule={rule}]  ({len(viols)} entries)")
        for v in viols:
            lines.append(f"  row {v.row:3d}: {v.message}")
        lines.append("")

    lines.append("=== SUMMARY ===")
    lines.append(f"Hard violations: {n_hard}")
    lines.append(f"Soft warnings:   {n_soft}")
    lines.append(f"Verdict:         {verdict}")
    lines.append("")

    log_path.write_text('\n'.join(lines), encoding='utf-8')


# ============================================================
# Loader
# ============================================================


def _read_rows(path: Path) -> list[_RowRecord]:
    """Read the XLSX and return one `_RowRecord` per non-empty data row.
    Performs AL1 parse and address normalization. No validation here."""
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    try:
        ws = wb.active
        records: list[_RowRecord] = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None:
                continue
            cells = list(row) + [None] * (6 - len(row))
            name, al1, al2, city, state, pc = cells[:6]

            if name is None and al1 is None and al2 is None:
                continue  # truly empty row — skip

            raw = {
                'Name': name, 'AddressLine1': al1, 'AddressLine2': al2,
                'City': city, 'State': state, 'Postcode': pc,
            }

            shipto_name, cust_num, al1_extra, al1_ok = parse_al1(al1)

            # Build address only when CSZ is present (else it's a
            # required-fields violation and address can't be keyed).
            city_n = _norm(city)
            state_n = _norm(state)
            pc_n = _fmt_postcode(pc)
            if city_n and state_n and pc_n and al2 is not None:
                addr, used_fallback = _normalize_address_with_status(al2, city, state, pc)
            else:
                addr = None
                used_fallback = False

            records.append(_RowRecord(
                row=row_idx, raw=raw,
                shipto_name=shipto_name, customer_number=cust_num,
                al1_extra=al1_extra, al1_parsed_ok=al1_ok,
                address=addr, address_used_fallback=used_fallback,
            ))
    finally:
        wb.close()
    return records


def _validate_and_report(
    records: list[_RowRecord],
    *,
    source_path: Path,
    strict: bool,
    log_dir_path: Optional[Path],
) -> list[_RowRecord]:
    """Run validation, write the report (if log_dir given), raise on
    strict-mode hard violations. Returns the records that pass the
    required-fields rule (suitable for downstream indexing)."""
    violations = validate_master(records)
    n_hard = sum(1 for v in violations if v.severity == 'hard')
    n_soft = sum(1 for v in violations if v.severity == 'soft')

    aborting = strict and n_hard > 0

    if log_dir_path is not None:
        log_path = log_dir_path / 'customer_master_validation.log'
        write_validation_report(
            violations,
            source_path=source_path,
            total_rows=len(records),
            log_path=log_path,
            strict=strict,
            aborted=aborting,
        )
        if n_hard > 0 or n_soft > 0:
            logger.warning(
                "customer-master validation: %d hard, %d soft (see %s)",
                n_hard, n_soft, log_path,
            )

    if aborting:
        raise MasterValidationError(
            f"customer master has {n_hard} hard validation violation(s) — "
            f"see {log_dir_path / 'customer_master_validation.log' if log_dir_path else '<log not written>'}"
        )

    rows_with_required_violation: set[int] = {
        v.row for v in violations if v.rule == RULE_REQUIRED_FIELDS
    }
    return [
        r for r in records
        if r.row not in rows_with_required_violation and r.address is not None
    ]


def load_address_to_customers(
    xlsx_path: Union[str, Path, None] = None,
    *,
    strict: bool = False,
    log_dir: Union[str, Path, None] = None,
) -> dict:
    """Load the customer-address XLSX and return
    `dict[NormalizedAddress, list[CustomerEntry]]`.

    Legacy entry point. New code should prefer `load_master` which
    returns a richer `CustomerMaster` with both address and
    customer-name indexes.
    """
    path = Path(xlsx_path) if xlsx_path else DEFAULT_ADDRESS_FILE
    log_dir_path = Path(log_dir) if log_dir is not None else None

    records = _read_rows(path)
    valid = _validate_and_report(
        records, source_path=path, strict=strict, log_dir_path=log_dir_path,
    )

    result: dict = {}
    for r in valid:
        entry = CustomerEntry(
            name=str(r.raw['Name']).strip() if r.raw['Name'] is not None else '',
            shipto_name=r.shipto_name,
        )
        result.setdefault(r.address, []).append(entry)

    logger.info(
        "loaded customer master (legacy dict): %d unique addresses from %d rows",
        len(result), len(records),
    )
    return result


def load_master(
    xlsx_path: Union[str, Path, None] = None,
    *,
    strict: bool = False,
    log_dir: Union[str, Path, None] = None,
) -> 'CustomerMaster':
    """Load the customer-master XLSX as a `CustomerMaster` instance with
    both address and customer-name indexes.

    Validation runs at load time. When `log_dir` is provided, a
    human-readable report is written to
    `<log_dir>/customer_master_validation.log`. When `strict=True`, any
    hard-rule violation aborts the load via `MasterValidationError`.
    Rows with required-fields violations are excluded.
    """
    path = Path(xlsx_path) if xlsx_path else DEFAULT_ADDRESS_FILE
    log_dir_path = Path(log_dir) if log_dir is not None else None

    records = _read_rows(path)
    valid = _validate_and_report(
        records, source_path=path, strict=strict, log_dir_path=log_dir_path,
    )

    entries = [
        MasterEntry(
            customer_name=str(r.raw['Name']).strip() if r.raw['Name'] is not None else '',
            shipto_name=r.shipto_name,
            address=r.address,
            customer_number=r.customer_number,
            row=r.row,
        )
        for r in valid
    ]
    master = CustomerMaster(entries)
    logger.info(
        "loaded customer master: %d entries / %d unique 4-tuple addresses / "
        "%d unique customer names",
        len(master.entries),
        len(master.by_address_4tuple),
        len(master.by_customer_name),
    )
    return master


_LEADING_NUMBER_RE = re.compile(r'^\s*(\d+(?:-\d+)?)')


def _leading_street_number(street: str):
    """Return the leading street number (e.g., '1176', '27072', '5980') or
    None if the street doesn't start with a digit. Treated as a hard
    discriminator in fuzzy fallback — different number = different address."""
    if not street:
        return None
    m = _LEADING_NUMBER_RE.match(street)
    return m.group(1) if m else None


def _normalize_name(name: str) -> str:
    """Lowercase, replace non-alphanumerics with spaces, collapse whitespace.

    Mirrors the legacy customer_lookup._normalize so the 0.88 difflib-ratio
    threshold stays semantically calibrated.
    """
    if not name:
        return ''
    s = re.sub(r'[^a-z0-9]+', ' ', str(name).lower())
    return re.sub(r'\s+', ' ', s).strip()


def _build_pairs_at_address(addr: NormalizedAddress, entries) -> list:
    """Helper: explode a single NormalizedAddress + entry list into pair tuples."""
    return [(addr, e) for e in entries]


def lookup_by_address(
    address_map: dict,
    street,
    city,
    state,
    postcode,
    *,
    fuzzy_threshold: int = DEFAULT_FUZZY_THRESHOLD,
) -> LookupResult:
    """Look up customer entries by address.

    Match strategy:
      1. Exact match on the normalized Address key.
      2. On miss: narrow to candidates sharing the same (city, state, postcode).
         Filter further to candidates with the same leading street number
         (a different number is always a different address). Score remaining
         candidates with rapidfuzz.fuzz.token_set_ratio over the street string.
         Accept the highest-scoring candidate at or above `fuzzy_threshold`.

    On `no_match`, `pairs` carries the best near-miss candidate's entries
    so the caller can show the closest miss.
    """
    key = _normalize_address(street, city, state, postcode)

    exact = address_map.get(key, [])
    if exact:
        return LookupResult(
            _build_pairs_at_address(key, exact),
            'address_exact', 100, 'n/a', 0,
        )

    csz = (key.city, key.state, key.postcode)
    bucket = [(a, es) for a, es in address_map.items()
              if (a.city, a.state, a.postcode) == csz]
    if not bucket:
        return LookupResult([], 'no_match', 0, 'n/a', 0)

    key_num = _leading_street_number(key.street)
    if key_num is not None:
        same_num = [(a, es) for a, es in bucket
                    if _leading_street_number(a.street) == key_num]
        if not same_num:
            # Hard fail: different street numbers ⇒ different address. Surface the
            # CSZ-bucket candidates as near-miss diagnostic.
            near_miss_pairs = [(a, e) for a, es in bucket for e in es]
            return LookupResult(near_miss_pairs, 'no_match', 0, 'n/a', 0)
        bucket = same_num

    best_addr, best_entries, best_score = None, [], 0
    for cand_addr, cand_entries in bucket:
        score = fuzz.token_set_ratio(key.street, cand_addr.street)
        if score > best_score:
            best_score, best_addr, best_entries = score, cand_addr, cand_entries

    if best_score >= fuzzy_threshold:
        return LookupResult(
            _build_pairs_at_address(best_addr, best_entries),
            'address_fuzzy', int(best_score), 'n/a', 0,
        )
    near_miss_pairs = (
        _build_pairs_at_address(best_addr, best_entries) if best_addr else []
    )
    return LookupResult(near_miss_pairs, 'no_match', int(best_score), 'n/a', 0)


def _disambiguate_by_name(
    pairs: list,
    consignee_name,
    consignee_norm: str,
    name_fuzzy_threshold: int,
) -> tuple:
    """When address lookup returns multiple candidates at the same address,
    try to narrow to one by matching the PDF consignee name against each
    candidate's `CustomerEntry.name`. Same 4-tier match the name-fallback
    uses, but scoped to just these N pairs.

    Returns (narrowed_pairs, name_method, name_score):
      - narrowed_pairs: list with exactly one (Address, CustomerEntry) on
        success, empty on `tried_failed`.
      - name_method: 'exact' | 'case_insensitive' | 'normalized' | 'fuzzy'
        | 'tried_failed'.
      - name_score: 100 for tiers 1–3, WRatio for tier 4, best WRatio seen on
        tried_failed (0 if no candidates).
    """
    # Tier 1 — exact verbatim
    hits = [(a, e) for a, e in pairs if e.name == consignee_name]
    if len(hits) == 1:
        return hits, 'exact', 100
    if len(hits) > 1:
        return [], 'tried_failed', 100  # multiple exact ties — ambiguous

    # Tier 2 — case-insensitive
    if consignee_name:
        lower = consignee_name.lower()
        hits = [(a, e) for a, e in pairs if e.name.lower() == lower]
        if len(hits) == 1:
            return hits, 'case_insensitive', 100
        if len(hits) > 1:
            return [], 'tried_failed', 100

    # Tier 3 — normalized form
    hits = [(a, e) for a, e in pairs
            if _normalize_name(e.name) == consignee_norm]
    if len(hits) == 1:
        return hits, 'normalized', 100
    if len(hits) > 1:
        return [], 'tried_failed', 100

    # Tier 4 — fuzzy WRatio. Take the top scorer if it clears the threshold
    # AND is not tied with another candidate.
    scored = [
        (fuzz.WRatio(consignee_norm, _normalize_name(e.name)), a, e)
        for a, e in pairs
    ]
    scored.sort(key=lambda x: -x[0])
    if not scored:
        return [], 'tried_failed', 0
    top_score = int(round(scored[0][0]))
    if scored[0][0] < name_fuzzy_threshold:
        return [], 'tried_failed', top_score
    if len(scored) > 1 and scored[1][0] >= scored[0][0]:
        return [], 'tried_failed', top_score  # tied at the top — still ambiguous
    return [(scored[0][1], scored[0][2])], 'fuzzy', top_score


def lookup_with_name_fallback(
    address_map: dict,
    consignee_name,
    street,
    city,
    state,
    postcode,
    *,
    fuzzy_threshold: int = DEFAULT_FUZZY_THRESHOLD,
    name_fuzzy_threshold: int = DEFAULT_NAME_FUZZY_THRESHOLD,
) -> LookupResult:
    """Look up customer entries by address first; on miss, fall back to a
    4-tier name match. When the consignee is a generic phrase like 'Federal
    Correctional Institution', skip the name-based steps (too generic to
    match safely).

    When address lookup returns multiple candidates at the same address,
    try to narrow them down by matching the consignee name against each
    candidate (same 4-tier match, scoped to that subset). If the name
    uniquely identifies one candidate, the result is `address_disambiguated_by_name`.

    Returns a LookupResult; see the LookupResult docstring for the meaning
    of each field. On no_match / multi_match_unresolved, `pairs` holds the
    best near-miss for diagnostic purposes.
    """
    addr_result = lookup_by_address(
        address_map, street, city, state, postcode,
        fuzzy_threshold=fuzzy_threshold,
    )

    consignee_norm = _normalize_name(consignee_name)

    if addr_result.cm_method in ('address_exact', 'address_fuzzy'):
        # Single match → done.
        if len(addr_result.pairs) <= 1:
            return addr_result

        # Multi-match. Generic consignee → can't disambiguate safely.
        if not consignee_norm or consignee_norm in _GENERIC_CONSIGNEE_NORMS:
            return LookupResult(
                addr_result.pairs, 'multi_match_unresolved',
                addr_result.addr_score, 'n/a', 0,
            )

        narrowed, name_method, name_score = _disambiguate_by_name(
            addr_result.pairs, consignee_name, consignee_norm,
            name_fuzzy_threshold,
        )
        if narrowed:
            return LookupResult(
                narrowed, 'address_disambiguated_by_name',
                addr_result.addr_score, name_method, name_score,
            )
        return LookupResult(
            addr_result.pairs, 'multi_match_unresolved',
            addr_result.addr_score, name_method, name_score,
        )

    # Address tier didn't match. Try name fallback unless consignee is generic.
    if not consignee_norm or consignee_norm in _GENERIC_CONSIGNEE_NORMS:
        return LookupResult(
            addr_result.pairs, 'no_match',
            addr_result.addr_score, 'n/a', 0,
        )

    found, name_method, name_score = _name_fallback_search(
        address_map, consignee_name, consignee_norm, name_fuzzy_threshold,
    )
    if name_method != 'tried_failed':
        return LookupResult(
            found, 'name_fallback',
            addr_result.addr_score, name_method, name_score,
        )

    # Both tiers missed. Surface whichever near-miss scored higher.
    if name_score > addr_result.addr_score:
        return LookupResult(
            found, 'no_match',
            addr_result.addr_score, 'tried_failed', name_score,
        )
    return LookupResult(
        addr_result.pairs, 'no_match',
        addr_result.addr_score, 'tried_failed', name_score,
    )


def _name_fallback_search(
    address_map: dict,
    consignee_name,
    consignee_norm: str,
    name_fuzzy_threshold: int,
) -> tuple:
    """Run the 4-tier name match against every CustomerEntry in the map.

    Returns (pairs, name_method, name_score) — same shape as
    `_disambiguate_by_name`, but searching the entire address map instead of
    a constrained candidate list. On a tier-1/2/3 hit, returns *all* pairs
    sharing the matched normalized name (one customer can have many
    addresses). On `tried_failed`, `pairs` is the best WRatio-near-miss for
    diagnostics.
    """
    by_norm: dict = {}
    canonical: dict = {}
    for addr, entries in address_map.items():
        for e in entries:
            n = _normalize_name(e.name)
            if not n:
                continue
            by_norm.setdefault(n, []).append((addr, e))
            canonical.setdefault(n, e.name)

    # Tier 1 — exact verbatim
    for addr, entries in address_map.items():
        for e in entries:
            if e.name == consignee_name:
                return by_norm[_normalize_name(e.name)], 'exact', 100

    # Tier 2 — case-insensitive
    if consignee_name:
        lower = consignee_name.lower()
        for n, pairs in by_norm.items():
            if canonical[n].lower() == lower:
                return pairs, 'case_insensitive', 100

    # Tier 3 — normalized
    if consignee_norm in by_norm:
        return by_norm[consignee_norm], 'normalized', 100

    # Tier 4 — fuzzy WRatio
    best_norm, best_score = None, 0.0
    for cand_norm in by_norm.keys():
        s = fuzz.WRatio(consignee_norm, cand_norm)
        if s > best_score:
            best_score, best_norm = s, cand_norm

    name_score = int(round(best_score))
    if best_score >= name_fuzzy_threshold:
        return by_norm[best_norm], 'fuzzy', name_score

    near_miss = by_norm.get(best_norm, []) if best_norm else []
    return near_miss, 'tried_failed', name_score


# ============================================================
# Stage-2 customer matching
# ============================================================
#
# Vocabulary:
#   "Customer name" = `Name` column (the billing entity, output of matching).
#   "Ship-to name"  = parsed AL1 prefix (used internally for row narrowing,
#                     never called the customer name).
#   Address match key = the 4-tuple (street, city, state, postcode).
#   line_2 is preserved on NormalizedAddress for display but is NOT in the
#   match key.


# Default minimum WRatio for the name-disambig matrix to accept a
# fuzzy-name hit when a multi-row 4-tuple address match needs to be
# narrowed. Tunable; user-confirmed starting value 75.
NAME_DISAMBIG_THRESHOLD = 75

# Customer-name deny-list. Matches against these rows are an error
# (rows 182, 183 in the master are 'Highland Beef Farms Inventory' —
# internal one-offs, never a real customer).
_HBF_INVENTORY_NORM = 'highland beef farms inventory'


@dataclass(frozen=True)
class MasterEntry:
    """One canonical row from the customer master.

    `customer_name` is the `Name` column (THE billing entity — what
    appears on the QuickBooks bill). `shipto_name` is the parsed AL1
    prefix (the destination's identity, used for matching but never
    called the customer name).
    """
    customer_name: str
    shipto_name: str
    address: NormalizedAddress
    customer_number: Optional[str]
    row: int                              # 1-indexed XLSX row, for diagnostics


class CustomerMaster:
    """The customer master with two public indexes:

      - `by_address_4tuple` (street, city, state, postcode) → entries.
        Primary for invoice matching. line_2 is NOT in the key.
      - `by_customer_name` normalized Customer Name (Name column) →
        entries. The public 'lookup by name' API.

    The master also exposes `entries` for callers that want to scan
    the flat list (e.g. tools).
    """

    def __init__(self, entries: list[MasterEntry]):
        self.entries: list[MasterEntry] = list(entries)
        self.by_address_4tuple: dict[
            tuple[str, str, str, str], list[MasterEntry]
        ] = {}
        self.by_customer_name: dict[str, list[MasterEntry]] = {}
        for e in self.entries:
            key4 = (e.address.street, e.address.city,
                    e.address.state, e.address.postcode)
            self.by_address_4tuple.setdefault(key4, []).append(e)
            cn_norm = _normalize_name(e.customer_name)
            if cn_norm:
                self.by_customer_name.setdefault(cn_norm, []).append(e)

    def lookup_address(
        self, addr_4tuple: tuple[str, str, str, str],
    ) -> list[MasterEntry]:
        """Return entries whose 4-tuple address (street, city, state,
        postcode) exactly matches. Empty list on miss."""
        return list(self.by_address_4tuple.get(tuple(addr_4tuple), []))

    def lookup_customer_name(self, name: str) -> list[MasterEntry]:
        """Return entries whose Customer Name (`Name` column) normalizes
        to the same form as `name`. The public by-name lookup. Empty
        list on miss."""
        return list(self.by_customer_name.get(_normalize_name(name), []))


# ---- Match results ---------------------------------------------------------


class MatchMethod:
    """String constants for SourceMatchResult.method and
    InvoiceMatchResult.method. Kept as plain str so they grep cleanly
    in logs."""
    NO_INPUT       = 'no_input'         # source had no usable address
    NO_MATCH       = 'no_match'         # 4-tuple lookup returned 0 rows
    UNIQUE         = 'unique'           # 4-tuple lookup returned exactly 1 row
    DISAMBIGUATED  = 'disambiguated'    # multi-row + name matrix picked one
    AMBIGUOUS      = 'ambiguous'        # multi-row + name matrix below threshold

    AGREE                 = 'agree'                  # both sources agreed
    BOL_WINS_DISAGREEMENT = 'bol_wins_disagreement'  # both resolved, disagreed; BOL won
    BOL_ONLY              = 'bol_only'               # only BOL resolved
    INV_ONLY              = 'inv_only'               # only page-1 resolved
    HARD_FAIL             = 'hard_fail'              # neither resolved
    DENIED                = 'denied'                 # matched a deny-listed row


@dataclass(frozen=True)
class SourceMatchResult:
    """Outcome of `run_match_for_source` against one extraction source."""
    method: str
    entry: Optional[MasterEntry] = None
    score: Optional[int] = None
    candidates: tuple[MasterEntry, ...] = ()
    matrix: tuple[tuple[str, str, int], ...] = ()  # (cand, master_shipto_name, score)


@dataclass(frozen=True)
class InvoiceMatchResult:
    """Final cross-source match for one invoice."""
    customer_name: Optional[str]              # the answer; None on hard fail / denied
    method: str
    bol: SourceMatchResult                    # never None (NO_INPUT if not run)
    inv: SourceMatchResult                    # never None (NO_INPUT if not run)
    severity: str                             # 'ok' | 'info' | 'severe'
    fail_reason: Optional[str] = None


# ---- Matcher implementation ------------------------------------------------


def _shipto_for(extraction):
    """Extract the inner ShipTo from either an InvoiceExtraction or a
    BolExtraction (or anything with a `.ship_to` attribute). Returns
    None if no usable extraction."""
    if extraction is None:
        return None
    return getattr(extraction, 'ship_to', None)


def run_match_for_source(extraction, master: CustomerMaster) -> SourceMatchResult:
    """Pure per-source matching. Caller passes ONE extraction (BOL or
    page-1); we look up its ShipTo's 4-tuple address against the
    master and, if multi-row, run the name-disambig matrix.

    Knows nothing about BOL vs page-1 priority — that's the
    orchestrator's concern.
    """
    ship_to = _shipto_for(extraction)
    if ship_to is None or ship_to.address is None or not ship_to.address.street:
        return SourceMatchResult(method=MatchMethod.NO_INPUT)

    addr = ship_to.address
    key4 = (addr.street, addr.city, addr.state, addr.postcode)
    rows = master.lookup_address(key4)

    if not rows:
        return SourceMatchResult(method=MatchMethod.NO_MATCH)

    if len(rows) == 1:
        return SourceMatchResult(
            method=MatchMethod.UNIQUE,
            entry=rows[0],
            candidates=(rows[0],),
        )

    # Multi-row → fuzzy name matrix over (cand × master_shipto_name).
    cands = list(ship_to.name_candidates) if ship_to.name_candidates else []
    if not cands and ship_to.name:
        cands = [ship_to.name]
    cands = [c for c in cands if c]  # drop empty strings

    matrix: list[tuple[str, str, int]] = []
    best_pair: Optional[tuple[str, MasterEntry]] = None
    best_score = 0
    for cand in cands:
        cand_norm = _normalize_name(cand)
        if not cand_norm:
            continue
        for row in rows:
            row_norm = _normalize_name(row.shipto_name)
            if not row_norm:
                continue
            score = int(round(fuzz.WRatio(cand_norm, row_norm)))
            matrix.append((cand, row.shipto_name, score))
            if score > best_score:
                best_score = score
                best_pair = (cand, row)

    if best_pair is not None and best_score >= NAME_DISAMBIG_THRESHOLD:
        return SourceMatchResult(
            method=MatchMethod.DISAMBIGUATED,
            entry=best_pair[1],
            score=best_score,
            candidates=tuple(rows),
            matrix=tuple(matrix),
        )

    return SourceMatchResult(
        method=MatchMethod.AMBIGUOUS,
        score=best_score,
        candidates=tuple(rows),
        matrix=tuple(matrix),
    )


def _is_resolved(r: SourceMatchResult) -> bool:
    return r.method in (MatchMethod.UNIQUE, MatchMethod.DISAMBIGUATED)


def _check_deny_list(result: InvoiceMatchResult) -> InvoiceMatchResult:
    """Post-match deny-list. If the matched customer is the
    Highland Beef Farms Inventory pseudo-customer (rows 182, 183 —
    internal one-offs, never a real billable customer), refuse the
    match."""
    if (result.customer_name and
            _normalize_name(result.customer_name) == _HBF_INVENTORY_NORM):
        return InvoiceMatchResult(
            customer_name=None,
            method=MatchMethod.DENIED,
            bol=result.bol, inv=result.inv,
            severity='severe',
            fail_reason=(
                f"matched HBF Inventory row ({result.customer_name!r}) — "
                f"internal one-off, not a real customer"
            ),
        )
    return result


def match_invoice_customer(
    invoice_extr,
    bol_extr,
    master: CustomerMaster,
) -> InvoiceMatchResult:
    """Stage-2 orchestrator. Runs both sources independently against
    the master, then picks the best.

    BOL is preferred when both sources resolve disagreeing customers
    (per user policy: 'BOL is highly preferred'). Hard fails if no
    source produced any address, or if neither source resolved a
    customer.
    """
    bol_result = run_match_for_source(bol_extr, master)
    inv_result = run_match_for_source(invoice_extr, master)

    # Phase 0: pre-flight. At least ONE source must produce an address.
    if (bol_result.method == MatchMethod.NO_INPUT
            and inv_result.method == MatchMethod.NO_INPUT):
        return InvoiceMatchResult(
            customer_name=None,
            method=MatchMethod.HARD_FAIL,
            bol=bol_result, inv=inv_result,
            severity='severe',
            fail_reason='no usable address from invoice',
        )

    bol_resolved = _is_resolved(bol_result)
    inv_resolved = _is_resolved(inv_result)

    if bol_resolved and inv_resolved:
        if bol_result.entry.customer_name == inv_result.entry.customer_name:
            return _check_deny_list(InvoiceMatchResult(
                customer_name=bol_result.entry.customer_name,
                method=MatchMethod.AGREE,
                bol=bol_result, inv=inv_result,
                severity='ok',
            ))
        # Disagreement: BOL wins per user policy. Severity depends on
        # whether both sources were UNIQUE (the strongest disagreement
        # signal — independent address locks landed on different rows)
        # vs. one of them having needed name disambig.
        if (bol_result.method == MatchMethod.UNIQUE
                and inv_result.method == MatchMethod.UNIQUE):
            severity = 'severe'
        else:
            severity = 'info'
        return _check_deny_list(InvoiceMatchResult(
            customer_name=bol_result.entry.customer_name,
            method=MatchMethod.BOL_WINS_DISAGREEMENT,
            bol=bol_result, inv=inv_result,
            severity=severity,
        ))

    if bol_resolved:
        return _check_deny_list(InvoiceMatchResult(
            customer_name=bol_result.entry.customer_name,
            method=MatchMethod.BOL_ONLY,
            bol=bol_result, inv=inv_result,
            severity='ok',
        ))

    if inv_resolved:
        return _check_deny_list(InvoiceMatchResult(
            customer_name=inv_result.entry.customer_name,
            method=MatchMethod.INV_ONLY,
            bol=bol_result, inv=inv_result,
            severity='ok',
        ))

    return InvoiceMatchResult(
        customer_name=None,
        method=MatchMethod.HARD_FAIL,
        bol=bol_result, inv=inv_result,
        severity='severe',
        fail_reason='neither BOL nor consignee resolved a customer',
    )


def format_match_log(result: InvoiceMatchResult) -> str:
    """Multi-line human-readable description of the match outcome.
    Suitable for a per-invoice log when the case is non-trivial
    (multi-row 4-tuple, BOL vs page-1 disagreement, etc.)."""
    lines = []
    lines.append(
        f"customer match: method={result.method}  severity={result.severity}"
    )
    if result.customer_name:
        lines.append(f"  customer_name: {result.customer_name!r}")
    if result.fail_reason:
        lines.append(f"  fail_reason:   {result.fail_reason}")

    for label, src in [('BOL', result.bol), ('Page-1', result.inv)]:
        lines.append(f"  --- {label} source ---")
        lines.append(f"    method: {src.method}")
        if src.entry is not None:
            e = src.entry
            lines.append(
                f"    matched row {e.row}: customer_name={e.customer_name!r}  "
                f"shipto_name={e.shipto_name!r}"
            )
            lines.append(
                f"      address: street={e.address.street!r} "
                f"city={e.address.city!r} state={e.address.state!r} "
                f"postcode={e.address.postcode!r}"
            )
        if src.score is not None:
            lines.append(f"    score: {src.score}  (threshold {NAME_DISAMBIG_THRESHOLD})")
        if len(src.candidates) > 1:
            lines.append(
                f"    candidate rows ({len(src.candidates)}): "
                + ', '.join(f'row {c.row} ({c.customer_name})' for c in src.candidates)
            )
        if src.matrix:
            lines.append(f"    name matrix (sorted by score, highest first):")
            for cand, shipto, score in sorted(src.matrix, key=lambda t: -t[2]):
                lines.append(
                    f"      WRatio({cand!r}, {shipto!r}) = {score}"
                )

    return '\n'.join(lines)
